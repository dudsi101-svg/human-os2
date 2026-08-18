from __future__ import annotations

import json

from fastapi import APIRouter, Depends, HTTPException
from fastapi.responses import JSONResponse, Response
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from sqlalchemy.orm import Session

from .. import consent_catalog
from ..authz import require_client_self
from ..db import get_db
from ..hos_bridge import ConsentService, record_event
from ..models import (
    AIUsageCounter,
    Challenge,
    ChallengeBlock,
    ChallengeEntry,
    ChallengeParticipant,
    ChallengeReport,
    CheckinRevision,
    CoachClientRelationship,
    ConsentRecord,
    ConsultSlot,
    DailyNutritionLog,
    Document,
    Goal,
    IdempotencyKey,
    Measurement,
    Message,
    MessageThread,
    Notification,
    NotificationPreference,
    NotificationSetting,
    NutritionPlan,
    NutritionPlanVersion,
    Observation,
    OcrTask,
    OnboardingAnswer,
    OnboardingSession,
    OnboardingSummaryItem,
    PaymentRecord,
    PaymentSchedule,
    ProfileField,
    ProgressPhoto,
    PushSubscription,
    Receipt,
    Reminder,
    ScheduleCompletion,
    ScheduleItem,
    StoredFile,
    TrainingPlan,
    TrainingPlanVersion,
    User,
    WeeklyCheckin,
    WorkoutEntry,
    WorkoutSession,
    now_iso,
)
from ..schemas import ConsentDeclineIn, ConsentGrantIn, DeletionRequestIn
from ..security import current_user, revoke_other_sessions, verify_password
from ..storage import storage

router = APIRouter(prefix="/api/me", tags=["privacy"])


def _rows(db: Session, model, **filters) -> list[dict]:
    out = []
    for row in db.query(model).filter_by(**filters).all():
        d = {c.name: getattr(row, c.name) for c in model.__table__.columns}
        for key in ("payload_json", "content_json"):
            if key in d and isinstance(d[key], str):
                try:
                    parsed = json.loads(d[key])
                except (ValueError, TypeError):
                    # Świadome zignorowanie: eksport danych (prawo do
                    # przenoszenia) musi się udać nawet przy uszkodzonym
                    # JSON-ie w pojedynczym rekordzie — pole zostaje w
                    # eksporcie w surowej postaci *_json zamiast znikać.
                    continue
                d[key.removesuffix("_json")] = parsed
                d.pop(key)
        out.append(d)
    return out


@router.get("/consents")
def my_consents(user: User = Depends(current_user), db: Session = Depends(get_db)):
    """Pełna historia zgód podmiotu + katalog kategorii (cel, zakres,
    odbiorcy, okres, dobrowolność, sposób wycofania, wersja dokumentu) —
    jedno źródło danych dla ekranu zgód i Profilu/Prywatności."""
    rows = (
        db.query(ConsentRecord)
        .filter(ConsentRecord.subject_id == user.id)
        .order_by(ConsentRecord.granted_at)
        .all()
    )
    grantee_names = {}
    for r in rows:
        if r.grantee_id == consent_catalog.SYSTEM_GRANTEE:
            grantee_names[r.grantee_id] = "Aplikacja Dzik OS"
            continue
        grantee = db.get(User, r.grantee_id)
        grantee_names[r.grantee_id] = grantee.display_name if grantee else r.grantee_id
    return {
        "document_version": consent_catalog.CONSENT_DOC_VERSION,
        "catalog": consent_catalog.catalog_payload(),
        "consents": [
            {
                "id": r.id,
                "grantee_id": r.grantee_id,
                "grantee_name": grantee_names.get(r.grantee_id),
                "category": r.category,
                "legal_basis": r.legal_basis,
                "source": r.source,
                "purpose": r.purpose,
                "domain": r.domain,
                "actions": r.actions,
                "allow_sensitive": r.allow_sensitive,
                "consent_text_version": r.consent_text_version,
                "document_version_current": (
                    r.consent_text_version == consent_catalog.CONSENT_DOC_VERSION
                ),
                "granted_at": r.granted_at,
                "expires_at": r.expires_at,
                "revoked_at": r.revoked_at,
                "confirmed_at": r.confirmed_at,
                "denied_at": r.denied_at,
            }
            for r in rows
        ],
    }


@router.post("/consents", status_code=201)
def grant_consent(
    body: ConsentGrantIn,
    user: User = Depends(current_user),
    db: Session = Depends(get_db),
):
    """Udzielenie zgody JEDNEJ kategorii. Brak jakiejkolwiek ścieżki
    „zaakceptuj wszystko" — każda kategoria to osobne wywołanie z osobnym
    wpisem w audycie."""
    cat = consent_catalog.category_by_key(body.category)
    if cat is None:
        raise HTTPException(status_code=422, detail="Nieznana kategoria zgody")
    if cat.grantee_kind == "COACH" and not body.grantee_id:
        raise HTTPException(status_code=422, detail="Wskaż odbiorcę (trenera)")
    row = ConsentService.grant_category(
        db,
        subject_id=user.id,
        category_key=body.category,
        grantee_id=body.grantee_id,
        actions=body.actions,
        source="SUBJECT",
        # Zgoda nadana osobiście przez podmiot jest potwierdzona z definicji.
        confirmed=True,
    )
    db.commit()
    return {"id": row.id, "category": row.category,
            "consent_text_version": row.consent_text_version}


@router.post("/consents/decline", status_code=201)
def decline_consent(
    body: ConsentDeclineIn,
    user: User = Depends(current_user),
    db: Session = Depends(get_db),
):
    """Jawna odmowa zgody OPCJONALNEJ — równie prosta jak udzielenie,
    zapisywana z pełną historią (wiersz z denied_at nigdy nie autoryzuje).
    Kategorii wymaganych (podstawa umowna) nie da się „odmówić" tą ścieżką
    — ich zakończenie to zakończenie współpracy/konta."""
    cat = consent_catalog.category_by_key(body.category)
    if cat is None:
        raise HTTPException(status_code=422, detail="Nieznana kategoria zgody")
    if cat.required:
        raise HTTPException(
            status_code=422,
            detail="Ta kategoria wynika z umowy — zakończenie przetwarzania "
            "to zakończenie współpracy albo usunięcie konta",
        )
    if cat.grantee_kind == "COACH" and not body.grantee_id:
        raise HTTPException(status_code=422, detail="Wskaż odbiorcę (trenera)")
    # Odmowa zamyka też ewentualną oczekującą deklarację z onboardingu.
    grantee = (
        consent_catalog.SYSTEM_GRANTEE
        if cat.grantee_kind == "SYSTEM"
        else body.grantee_id
    )
    for pending in (
        db.query(ConsentRecord)
        .filter(
            ConsentRecord.subject_id == user.id,
            ConsentRecord.category == cat.key,
            ConsentRecord.grantee_id == grantee,
            ConsentRecord.revoked_at.is_(None),
            ConsentRecord.denied_at.is_(None),
        )
        .all()
    ):
        ConsentService.revoke(db, consent_id=pending.id, subject_id=user.id)
    row = ConsentService.decline_category(
        db, subject_id=user.id, category_key=body.category, grantee_id=body.grantee_id
    )
    db.commit()
    return {"id": row.id, "category": row.category, "denied_at": row.denied_at}


@router.post("/consents/{consent_id}/confirm")
def confirm_consent(
    consent_id: str,
    user: User = Depends(current_user),
    db: Session = Depends(get_db),
):
    """Jawne potwierdzenie przez podmiot zgody zarejestrowanej przy
    onboardingu (deklaracja zebrana przez trenera)."""
    from ..models import ConsentRecord, now_iso

    row = db.get(ConsentRecord, consent_id)
    if row is None or row.subject_id != user.id or row.revoked_at is not None:
        raise HTTPException(status_code=404, detail="Nie znaleziono")
    if row.confirmed_at is None:
        row.confirmed_at = now_iso()
        record_event(
            db,
            action="CONSENT_CONFIRMED",
            actor_id=user.id,
            subject_ids=[user.id],
            payload={"consent_id": row.id, "grantee_id": row.grantee_id,
                     "purpose": row.purpose, "domain": row.domain},
            summary=f"Potwierdzenie zgody {row.purpose}/{row.domain} przez podmiot",
        )
        db.commit()
    return {"id": row.id, "confirmed_at": row.confirmed_at}


@router.post("/consents/{consent_id}/revoke")
def revoke_consent(
    consent_id: str,
    user: User = Depends(current_user),
    db: Session = Depends(get_db),
):
    try:
        row = ConsentService.revoke(db, consent_id=consent_id, subject_id=user.id)
    except PermissionError:
        raise HTTPException(status_code=404, detail="Nie znaleziono") from None
    if row.category == "przypomnienia":
        # Wycofanie zgody na przypomnienia natychmiast kończy przyszłe
        # przetwarzanie: wszystkie subskrypcje push znikają (żaden kanał
        # doręczeń nie zostaje).
        removed = (
            db.query(PushSubscription)
            .filter(PushSubscription.user_id == user.id)
            .delete()
        )
        if removed:
            record_event(
                db, action="PUSH_UNSUBSCRIBED", actor_id=user.id,
                subject_ids=[user.id],
                payload={"reason": "consent_revoked", "count": removed},
                summary="Usunięto subskrypcje push po wycofaniu zgody",
            )
    db.commit()
    return {"id": row.id, "revoked_at": row.revoked_at}


def _collect_export(db: Session, user: User) -> dict:
    client_id = user.id
    threads = (
        db.query(MessageThread)
        .filter((MessageThread.client_id == client_id) | (MessageThread.coach_id == client_id))
        .all()
    )
    messages = []
    for t in threads:
        messages.extend(_rows(db, Message, thread_id=t.id))
    plans = _rows(db, TrainingPlan, client_id=client_id)
    plan_versions = []
    for p in plans:
        plan_versions.extend(_rows(db, TrainingPlanVersion, plan_id=p["id"]))
    nplans = _rows(db, NutritionPlan, client_id=client_id)
    nversions = []
    for p in nplans:
        nversions.extend(_rows(db, NutritionPlanVersion, plan_id=p["id"]))
    checkins = _rows(db, WeeklyCheckin, client_id=client_id)
    revisions = []
    for c in checkins:
        revisions.extend(_rows(db, CheckinRevision, checkin_id=c["id"]))
    sessions = _rows(db, WorkoutSession, client_id=client_id)
    entries = []
    for s in sessions:
        entries.extend(_rows(db, WorkoutEntry, session_id=s["id"]))
    pay_schedules = _rows(db, PaymentSchedule, client_id=client_id)
    pay_records = []
    for s in pay_schedules:
        pay_records.extend(_rows(db, PaymentRecord, schedule_id=s["id"]))
    schedule_items = _rows(db, ScheduleItem, client_id=client_id)
    schedule_completions = []
    for i in schedule_items:
        schedule_completions.extend(_rows(db, ScheduleCompletion, schedule_item_id=i["id"]))
    consult_slots = _rows(db, ConsultSlot, client_id=client_id)
    # Subskrypcje push: bez kluczy kryptograficznych subskrypcji (p256dh/
    # auth to sekrety kanału doręczeń — minimalizacja: nie kopiujemy ich
    # do pliku eksportu; sam endpoint wystarcza do rozliczalności).
    push_subs = [
        {"id": p.id, "endpoint": p.endpoint, "created_at": p.created_at}
        for p in db.query(PushSubscription).filter_by(user_id=client_id).all()
    ]
    receipts = _rows(db, Receipt, subject_id=client_id)
    # Wyzwania: udziały użytkownika + jego własne wpisy wyników (cudze
    # dane z wyzwań NIE wchodzą do eksportu — minimalizacja).
    challenge_participations = _rows(db, ChallengeParticipant, user_id=client_id)
    challenge_entries = []
    for cp in challenge_participations:
        challenge_entries.extend(_rows(db, ChallengeEntry, participant_id=cp["id"]))
    # Rozmowa startowa (onboarding): sesje, WSZYSTKIE wersje odpowiedzi
    # (także poprawione — historia należy do klienta) i podsumowanie.
    onboarding_sessions = _rows(db, OnboardingSession, client_id=client_id)
    onboarding_answers = []
    onboarding_summary = []
    for onb in onboarding_sessions:
        onboarding_answers.extend(_rows(db, OnboardingAnswer, session_id=onb["id"]))
        onboarding_summary.extend(_rows(db, OnboardingSummaryItem, session_id=onb["id"]))
    return {
        "export_version": "1.5",
        "user": {
            "id": user.id, "email": user.email, "display_name": user.display_name,
            "identity_id": user.identity_id, "created_at": user.created_at,
        },
        "profile_fields": _rows(db, ProfileField, client_id=client_id),
        "goals": _rows(db, Goal, client_id=client_id),
        "training_plans": plans,
        "training_plan_versions": plan_versions,
        "workout_sessions": sessions,
        "workout_entries": entries,
        "nutrition_plans": nplans,
        "nutrition_plan_versions": nversions,
        "schedule_items": schedule_items,
        "schedule_completions": schedule_completions,
        "reminders": _rows(db, Reminder, client_id=client_id),
        "weekly_checkins": checkins,
        "checkin_revisions": revisions,
        "measurements": _rows(db, Measurement, client_id=client_id),
        "progress_photos": _rows(db, ProgressPhoto, client_id=client_id),
        "documents": _rows(db, Document, client_id=client_id),
        "files": _rows(db, StoredFile, owner_user_id=client_id),
        "messages": messages,
        "payment_schedules": pay_schedules,
        "payment_records": pay_records,
        "consents": _rows(db, ConsentRecord, subject_id=client_id),
        "observations": _rows(db, Observation, client_id=client_id),
        "daily_nutrition_logs": _rows(db, DailyNutritionLog, client_id=client_id),
        "consult_slots": consult_slots,
        "push_subscriptions": push_subs,
        "notifications": _rows(db, Notification, user_id=client_id),
        "notification_preferences": _rows(db, NotificationPreference, user_id=client_id),
        "notification_settings": _rows(db, NotificationSetting, user_id=client_id),
        "onboarding_sessions": onboarding_sessions,
        "onboarding_answers": onboarding_answers,
        "onboarding_summary_items": onboarding_summary,
        "ai_usage_counters": _rows(db, AIUsageCounter, user_id=client_id),
        # Zadania przepisywania tekstu ze zdjęcia: rozpoznany tekst i
        # propozycja pól to dane osobowe (bywa, że zdrowotne) — wchodzą do
        # eksportu tak samo jak treść raportów (prawo do przenoszenia).
        "ocr_tasks": _rows(db, OcrTask, owner_user_id=client_id),
        "audit_receipts": receipts,
        "challenge_participations": challenge_participations,
        "challenge_entries": challenge_entries,
    }


@router.get("/export")
def export_my_data(user: User = Depends(current_user), db: Session = Depends(get_db)):
    """Eksport wszystkich danych użytkownika (prawo do przenoszenia danych).
    Zwraca komplet danych jako JSON; pliki wymienione są z identyfikatorami
    do pobrania przez /api/files/{id}."""
    export = _collect_export(db, user)
    record_event(
        db,
        action="DATA_EXPORTED",
        actor_id=user.id,
        subject_ids=[user.id],
        payload={"export_version": export["export_version"], "format": "json"},
        summary="Eksport danych użytkownika (JSON)",
    )
    db.commit()
    return JSONResponse(
        content=json.loads(json.dumps(export, default=str)),
        headers={"Content-Disposition": 'attachment; filename="dzik-os-export.json"'},
    )


@router.get("/export.xlsx")
def export_my_data_excel(user: User = Depends(current_user), db: Session = Depends(get_db)):
    """Ten sam eksport co /export, w formacie arkusza kalkulacyjnego —
    jeden arkusz na tabelę, wygodny do przejrzenia w Excelu/LibreOffice."""
    export = _collect_export(db, user)
    wb = Workbook()
    wb.remove(wb.active)
    for name, rows in export.items():
        if name == "user" or not isinstance(rows, list):
            continue
        sheet_name = name[:31]  # limit Excela na długość nazwy arkusza
        ws = wb.create_sheet(sheet_name)
        if not rows:
            ws.append(["(brak danych)"])
            continue
        headers = sorted({k for row in rows for k in row})
        ws.append(headers)
        for row in rows:
            ws.append([
                json.dumps(row[h], ensure_ascii=False, default=str)
                if isinstance(row.get(h), (dict, list))
                else row.get(h)
                for h in headers
            ])
        for idx, h in enumerate(headers, start=1):
            width = max(10, min(40, len(h) + 2))
            ws.column_dimensions[get_column_letter(idx)].width = width
    info = wb.create_sheet("konto", 0)
    info.append(["pole", "wartość"])
    for k, v in export["user"].items():
        info.append([k, v])

    from io import BytesIO

    buf = BytesIO()
    wb.save(buf)
    record_event(
        db,
        action="DATA_EXPORTED",
        actor_id=user.id,
        subject_ids=[user.id],
        payload={"export_version": export["export_version"], "format": "xlsx"},
        summary="Eksport danych użytkownika (Excel)",
    )
    db.commit()
    return Response(
        content=buf.getvalue(),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": 'attachment; filename="dzik-os-export.xlsx"'},
    )


@router.post("/deletion-request")
def request_deletion(
    body: DeletionRequestIn,
    user: User = Depends(current_user),
    db: Session = Depends(get_db),
):
    """Usunięcie danych: anonimizacja konta i danych osobowych oraz
    fizyczne usunięcie plików. Zapisy audytowe (łańcuch zdarzeń) pozostają —
    zawierają wyłącznie identyfikatory, nie dane zdrowotne w postaci jawnej.
    Dane rozliczeniowe (ewidencja płatności: kwoty, terminy, statusy)
    pozostają bez treści opisowych — obowiązek podatkowy administratora
    (art. 6 ust. 1 lit. c RODO). Operacja jest nieodwracalna i wymaga
    hasła + frazy potwierdzającej."""
    if not verify_password(body.password, user.password_hash):
        raise HTTPException(status_code=403, detail="Nieprawidłowe hasło")
    client_id = require_client_self(db, user)

    for stored in db.query(StoredFile).filter(StoredFile.owner_user_id == client_id).all():
        storage.delete(stored)
        stored.deleted_at = stored.deleted_at or "deleted"
        stored.filename = "usunięto"
        stored.storage_path = ""
    for f in db.query(ProfileField).filter(ProfileField.client_id == client_id).all():
        f.value = "[usunięto]"
    for m in db.query(Measurement).filter(Measurement.client_id == client_id).all():
        db.delete(m)
    for p in db.query(ProgressPhoto).filter(ProgressPhoto.client_id == client_id).all():
        db.delete(p)
    for o in db.query(Observation).filter(Observation.client_id == client_id).all():
        db.delete(o)
    for n in db.query(DailyNutritionLog).filter(DailyNutritionLog.client_id == client_id).all():
        db.delete(n)
    for sc in db.query(ScheduleCompletion).filter(ScheduleCompletion.client_id == client_id).all():
        sc.note = None
    for c in db.query(WeeklyCheckin).filter(WeeklyCheckin.client_id == client_id).all():
        c.payload_json = "{}"
        c.coach_response = None
    # Wolne teksty mogące zawierać dane osobowe/zdrowotne — anonimizacja
    # (liczby i daty zostają jako dane spseudonimizowane bez powiązania
    # z osobą po anonimizacji konta).
    for g in db.query(Goal).filter(Goal.client_id == client_id).all():
        g.title = "[usunięto]"
        g.description = None
    for rm in db.query(Reminder).filter(Reminder.client_id == client_id).all():
        rm.text = "[usunięto]"
    for si in db.query(ScheduleItem).filter(ScheduleItem.client_id == client_id).all():
        si.name = "[usunięto]"
        si.instruction = None
        si.author_note = None
        si.status = "ENDED"
    for ws in db.query(WorkoutSession).filter(WorkoutSession.client_id == client_id).all():
        ws.comment = None
        ws.pain_note = None
        for we in db.query(WorkoutEntry).filter(WorkoutEntry.session_id == ws.id).all():
            we.comment = None
    for doc in db.query(Document).filter(Document.client_id == client_id).all():
        doc.title = "[usunięto]"
        doc.status = "ARCHIVED"
        # Tekst przepisany ze skanu bywa daną zdrowotną (np. wynik badań) —
        # znika razem z resztą treści.
        doc.ocr_text = None
        doc.ocr_engine = None
        doc.ocr_at = None
    # Zadania przepisywania tekstu ze zdjęcia znikają w całości: niosą
    # rozpoznaną treść, a nie tylko metadane (ślad operacji zostaje w
    # niemutowalnym łańcuchu audytu, bez treści).
    db.query(OcrTask).filter(
        (OcrTask.owner_user_id == client_id) | (OcrTask.created_by == client_id)
    ).delete()
    # Wyzwania: trwałe wycofanie wszystkich udziałów — wpisy wyników są
    # usuwane, pseudonimy anonimizowane, agregaty grup oznaczone jako
    # skorygowane; wyzwania indywidualne (organizowane przez klienta) są
    # anulowane z anonimizacją treści. Zgłoszenia autorstwa użytkownika
    # tracą treść (mogła zawierać dane osobowe).
    for cp in db.query(ChallengeParticipant).filter_by(user_id=client_id).all():
        deleted_entries = (
            db.query(ChallengeEntry).filter_by(participant_id=cp.id).delete()
        )
        cp.status = "WITHDRAWN"
        cp.withdrawn_at = now_iso()
        cp.alias = None
        cp.share_result = False
        cp.ranking_opt_in = False
        cp.auto_count_workouts = False
        if deleted_entries:
            ch = db.get(Challenge, cp.challenge_id)
            if ch is not None:
                ch.aggregates_adjusted = True
    for ch in db.query(Challenge).filter_by(organizer_id=client_id).all():
        ch.title = "[usunięto]"
        ch.description = None
        if ch.status in ("DRAFT", "ACTIVE"):
            ch.status = "CANCELLED"
            ch.cancelled_at = now_iso()
    for rep in db.query(ChallengeReport).filter_by(reporter_id=client_id).all():
        rep.reason = "[usunięto]"
    db.query(ChallengeBlock).filter(
        (ChallengeBlock.blocker_id == client_id)
        | (ChallengeBlock.blocked_id == client_id)
    ).delete()
    # Subskrypcje push znikają w całości (kanał doręczeń przestaje istnieć).
    db.query(PushSubscription).filter(PushSubscription.user_id == client_id).delete()
    # Powiadomienia (treść centrum może zawierać nazwy z harmonogramu itp.),
    # preferencje i ustawienia doręczeń znikają razem z kontem.
    db.query(Notification).filter(Notification.user_id == client_id).delete()
    db.query(NotificationPreference).filter(
        NotificationPreference.user_id == client_id
    ).delete()
    db.query(NotificationSetting).filter(
        NotificationSetting.user_id == client_id
    ).delete()
    # Rozmowa startowa: odpowiedzi i podsumowanie zawierają wolny tekst
    # klienta (w tym dane zdrowotne) — znikają w całości razem z kontem,
    # tak jak treść raportów. Liczniki kosztów AI (same liczby) również.
    for onb in db.query(OnboardingSession).filter_by(client_id=client_id).all():
        db.query(OnboardingAnswer).filter_by(session_id=onb.id).delete()
        db.query(OnboardingSummaryItem).filter_by(session_id=onb.id).delete()
        onb.status = "ABANDONED"
        onb.abandoned_at = now_iso()
        onb.current_step_id = None
        onb.summary_mode_reason = None
    db.query(AIUsageCounter).filter(AIUsageCounter.user_id == client_id).delete()
    # Klucze idempotencji (metadane operacyjne z identyfikatorami zapisów)
    # znikają razem z kontem.
    db.query(IdempotencyKey).filter(IdempotencyKey.user_id == client_id).delete()
    # Konsultacje: odpięcie klienta od slotów; przyszłe rezerwacje wracają
    # do puli trenera jako wolne.
    for slot in db.query(ConsultSlot).filter(ConsultSlot.client_id == client_id).all():
        slot.client_id = None
        slot.booked_at = None
        if slot.status == "BOOKED":
            slot.status = "OPEN"
    # Ewidencja płatności: kwoty/terminy/statusy zostają (rozliczenia),
    # treści opisowe są usuwane.
    for ps in db.query(PaymentSchedule).filter(PaymentSchedule.client_id == client_id).all():
        for pr in db.query(PaymentRecord).filter(PaymentRecord.schedule_id == ps.id).all():
            pr.note = None
    for r in (
        db.query(CheckinRevision)
        .join(WeeklyCheckin, CheckinRevision.checkin_id == WeeklyCheckin.id)
        .filter(WeeklyCheckin.client_id == client_id)
        .all()
    ):
        r.payload_json = "{}"
    threads = db.query(MessageThread).filter(MessageThread.client_id == client_id).all()
    for t in threads:
        for m in db.query(Message).filter(Message.thread_id == t.id).all():
            m.body = "[usunięto]"
            m.file_id = None
    # Domknięcie dostępu: współprace kończą się, aktywne zgody zostają
    # cofnięte (trener traci dostęp także do pozostałych, zanonimizowanych
    # rekordów), a wszystkie sesje są unieważnione — żaden stary token nie
    # działa po usunięciu konta.
    for rel in (
        db.query(CoachClientRelationship)
        .filter(CoachClientRelationship.client_id == client_id)
        .all()
    ):
        if rel.status != "ENDED":
            rel.status = "ENDED"
            rel.ended_at = now_iso()
    for consent in (
        db.query(ConsentRecord)
        .filter(ConsentRecord.subject_id == client_id, ConsentRecord.revoked_at.is_(None))
        .all()
    ):
        consent.revoked_at = now_iso()
    revoke_other_sessions(db, client_id, keep_token=None)
    user.email = f"deleted-{user.id.lower()}@example.invalid"
    user.display_name = "Konto usunięte"
    user.status = "DELETED"
    user.anonymized_at = now_iso()
    record_event(
        db,
        action="ACCOUNT_ANONYMIZED",
        actor_id=user.id,
        subject_ids=[user.id],
        payload={"user_id": user.id},
        summary="Anonimizacja konta i usunięcie danych na żądanie użytkownika",
    )
    db.commit()
    return {"ok": True, "status": "DELETED"}
