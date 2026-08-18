"""Import bazy danych trenera z pliku arkusza (CSV / XLSX).

DWIE BAZY, JEDEN MECHANIZM:

* **ćwiczenia** — jeden wiersz = jedno ćwiczenie (`import_exercises_sheet`),
* **szablony treningowe** — jeden wiersz = jedno ćwiczenie w jednym dniu
  jednego szablonu (`import_templates_sheet`).

Różni je tylko zestaw kolumn i to, co powstaje na końcu; wszystko inne —
czytanie pliku, raport, próba przed zapisem, izolacja katalogu trenera —
jest wspólne.

CZTERY ZASADY (te same, co przy `import_exercises.py`):

1. **Próba przed zapisem.** Domyślnie `dry_run=True`: przelicza cały plik
   i zwraca dokładnie ten sam raport, ale nie zapisuje ani jednego wiersza.
   Zapis następuje dopiero po świadomym kliknięciu trenera.
2. **Nie zgadujemy.** Wartość spoza słownika NIE trafia do bazy: albo
   wiersz jest pomijany z opisem przyczyny (pola wymagane), albo pole
   zostaje puste i ląduje w raporcie (mięśnie). Nigdy „najbliższy klucz”.
3. **Praca trenera jest nienaruszalna.** W trybie `UZUPELNIJ` (domyślnym)
   import wypełnia wyłącznie PUSTE pola istniejącego ćwiczenia. Tryb
   `ZASTAP` nadpisuje, ale tylko wartościami niepustymi — pusta komórka
   nigdy nie kasuje tego, co już jest w bazie. Historia szablonów jest
   niemutowalna: import na istniejącym szablonie tworzy NOWĄ WERSJĘ.
4. **Import jest idempotentny.** Drugi przebieg tego samego pliku kończy
   się zerem utworzonych i zerem zmienionych pozycji; szablon o
   niezmienionej treści nie dostaje pustej wersji „bo import”.

Błąd pojedynczego wiersza nie przerywa importu — trafia do `errors` i
lecimy dalej. Import zawsze idzie do katalogu KONKRETNEGO trenera
(`coach_id`); katalogi trenerów są rozłączne.

Granica roli (Human OS): moduł niczego nie ocenia i nie dobiera — przenosi
materiał trenera do jego własnej bazy. Wybór ćwiczeń i kształt planu
pozostają decyzją człowieka.
"""

from __future__ import annotations

import csv
import io
import json
from dataclasses import dataclass, field
from typing import Any

from sqlalchemy.orm import Session

from .exercise_parser import SOURCE_IMPORTED, map_muscle_phrase
from .models import (
    Exercise,
    ImportSnapshot,
    TrainingPlan,
    TrainingPlanVersion,
    new_id,
    now_iso,
)
from .muscles import (
    EXERCISE_LEVELS,
    LEVEL_LABELS,
    MOVEMENT_PATTERNS,
    MUSCLE_GROUPS,
    MUSCLE_LABELS,
    PATTERN_LABELS,
    fold,
    join_muscles,
)

#: Twardy limit rozmiaru pliku. 5 MB to z zapasem kilkanaście tysięcy
#: wierszy CSV — a jednocześnie granica, przy której czytanie w pamięci
#: pozostaje bezpieczne dla maszyny 1 GB (patrz fly.toml).
MAX_BYTES = 5 * 1024 * 1024
#: Limit wierszy danych (bez nagłówka). Reszta pliku trafia do raportu
#: jako pominięta — cicha utrata części pliku byłaby gorsza niż odmowa.
MAX_ROWS = 2000
#: Separator elementów listy WEWNĄTRZ jednej komórki (kroki, błędy, tagi).
#: Pionowa kreska, bo przecinek i średnik są zajęte przez sam format CSV.
LIST_SEPARATOR = "|"
#: Separator listy mięśni w komórce — tu przecinek jest naturalny i
#: bezpieczny, bo nazwy partii go nie zawierają.
MUSCLE_SEPARATOR = ","

XLSX_SUFFIXES = (".xlsx", ".xlsm")
CSV_SUFFIXES = (".csv", ".txt", ".tsv")


class SheetError(Exception):
    """Plik jest nie do przeczytania jako całość (złe kodowanie, brak
    wymaganych kolumn, pusty arkusz). Router zamienia to na 422.

    Odróżnione od błędu wiersza: błąd wiersza nie przerywa importu."""


# --- Kontrakt kolumn ---------------------------------------------------

@dataclass(frozen=True)
class Column:
    """Jedna kolumna pliku: nazwa kanoniczna, wymagalność, opis i przykład.

    `aliases` to JAWNIE dopuszczone inne nagłówki tej samej kolumny — nie
    domysł, tylko wypisany kontrakt (angielska nazwa, wariant bez polskich
    znaków). Nagłówki spoza kontraktu trafiają do `unknown_columns`."""

    key: str
    label: str
    required: bool = False
    example: str = ""
    aliases: tuple[str, ...] = ()


EXERCISE_COLUMNS: tuple[Column, ...] = (
    Column("nazwa", "Nazwa ćwiczenia", required=True,
           example="Wyciskanie sztangi na ławce płaskiej", aliases=("name", "cwiczenie")),
    Column("grupa", "Grupa mięśniowa (słownik)", required=True,
           example="KLATKA", aliases=("kategoria", "muscle_group")),
    Column("opis", "Opis wykonania (technika)", required=True,
           example="Połóż się na ławce, chwyt nieco szerszy niż barki...",
           aliases=("how_to", "wykonanie", "technika")),
    Column("nazwa_en", "Nazwa angielska", example="Barbell bench press",
           aliases=("name_en", "nazwa_ang")),
    Column("efekt", "Co daje to ćwiczenie", example="Siła i masa klatki piersiowej",
           aliases=("benefit", "korzysc")),
    Column("sprzet", "Potrzebny sprzęt", example="sztanga, ławka płaska",
           aliases=("equipment", "wyposazenie")),
    Column("poziom", "Poziom trudności (słownik)", example="SREDNIOZAAWANSOWANY",
           aliases=("level", "trudnosc")),
    Column("wzorzec", "Wzorzec ruchu (słownik)", example="WYPYCHANIE_POZIOME",
           aliases=("pattern", "wzorzec_ruchu")),
    Column("miesnie_glowne", "Mięśnie główne (słownik, po przecinku)",
           example="KLATKA_PIERSIOWA", aliases=("muscles_primary", "miesnie_pierwszorzedowe")),
    Column("miesnie_pomocnicze", "Mięśnie pomocnicze (słownik, po przecinku)",
           example="TRICEPS,BARK_PRZEDNI", aliases=("muscles_secondary", "miesnie_wspomagajace")),
    Column("kroki", f"Kroki techniki, rozdzielone „{LIST_SEPARATOR}”",
           example="Ustaw łopatki|Opuść sztangę do klatki|Wypchnij",
           aliases=("steps", "kroki_techniki")),
    Column("bledy", f"Najczęstsze błędy, rozdzielone „{LIST_SEPARATOR}”",
           example="Odbijanie od klatki|Uniesione barki", aliases=("mistakes", "bledy_techniczne")),
    Column("wskazowki", f"Wskazówki trenerskie, rozdzielone „{LIST_SEPARATOR}”",
           example="Wbij stopy w podłogę|Łokcie pod kątem 45°", aliases=("cues", "cue")),
    Column("bezpieczenstwo", "Uwagi bezpieczeństwa",
           example="Przy bólu barku przerwij i skonsultuj się ze specjalistą.",
           aliases=("safety", "uwagi_bezpieczenstwa")),
    Column("latwiej", "Wersja łatwiejsza", example="Wyciskanie hantlami",
           aliases=("easier", "regresja")),
    Column("trudniej", "Wersja trudniejsza", example="Wyciskanie z pauzą",
           aliases=("harder", "progresja")),
    Column("tempo", "Sugerowane tempo", example="3-1-1-0",
           aliases=("tempo_hint", "tempo_ruchu")),
    Column("oddech", "Sposób oddychania", example="Wdech przy opuszczaniu, wydech przy wypychaniu",
           aliases=("breathing",)),
    Column("tagi", f"Etykiety, rozdzielone „{LIST_SEPARATOR}”",
           example="wielostawowe|siłowe", aliases=("tags", "etykiety")),
    Column("wideo_url", "Link do nagrania", example="https://...",
           aliases=("video_url", "wideo", "film")),
    Column("zrodlo", "Skąd pochodzi pozycja (proweniencja)",
           example="Biblioteka własna, 2026-08", aliases=("source_ref", "source")),
)

TEMPLATE_COLUMNS: tuple[Column, ...] = (
    Column("szablon", "Nazwa szablonu (grupuje wiersze)", required=True,
           example="FBW 3x w tygodniu", aliases=("template", "plan", "nazwa_szablonu")),
    Column("dzien", "Nazwa dnia treningowego", required=True,
           example="Dzień A — całe ciało", aliases=("day", "trening", "nazwa_dnia")),
    Column("cwiczenie", "Nazwa ćwiczenia", required=True,
           example="Przysiad ze sztangą z tyłu", aliases=("exercise", "nazwa")),
    Column("dzien_nr", "Kolejność dnia w szablonie (liczba)", example="1",
           aliases=("day_no", "nr_dnia", "kolejnosc_dnia")),
    Column("dzien_tygodnia", "Dzień tygodnia 1–7 (1 = poniedziałek)", example="1",
           aliases=("weekday", "dzien_tyg")),
    Column("pozycja", "Kolejność ćwiczenia w dniu (liczba)", example="1",
           aliases=("order", "lp", "nr")),
    Column("serie", "Liczba serii (tekst — dopuszczalne „3–4”)", example="4",
           aliases=("sets",)),
    Column("powtorzenia", "Powtórzenia (tekst — dopuszczalne „8–10”)", example="8-10",
           aliases=("reps", "powt")),
    Column("ciezar", "Obciążenie (tekst — dopuszczalne „RPE 8”, „bw”)", example="RPE 8",
           aliases=("weight", "obciazenie")),
    Column("tempo", "Tempo ruchu", example="3-1-1-0", aliases=("tempo_ruchu",)),
    Column("przerwa", "Przerwa między seriami", example="120 s",
           aliases=("rest", "odpoczynek")),
    Column("komentarz", "Uwaga trenera do pozycji", example="Ostatnia seria do 2 powtórzeń zapasu",
           aliases=("comment", "uwagi", "notatka")),
    Column("wideo_url", "Link do nagrania dla tej pozycji", example="https://...",
           aliases=("video_url", "wideo", "film")),
)


def _index(columns: tuple[Column, ...]) -> dict[str, str]:
    """Nagłówek (kanoniczny lub alias) → kanoniczna nazwa kolumny."""
    mapping: dict[str, str] = {}
    for column in columns:
        mapping[column.key] = column.key
        for alias in column.aliases:
            mapping[alias] = column.key
    return mapping


EXERCISE_HEADERS = _index(EXERCISE_COLUMNS)
TEMPLATE_HEADERS = _index(TEMPLATE_COLUMNS)


def _assert_columns() -> None:
    """Kontrakt kolumn nie może mieć dwuznacznych nagłówków ani wskazywać
    słowników spoza `muscles.py` — sprawdzamy to przy imporcie modułu,
    a nie dopiero na pliku trenera."""
    for columns, name in ((EXERCISE_COLUMNS, "ćwiczeń"), (TEMPLATE_COLUMNS, "szablonów")):
        seen: set[str] = set()
        for column in columns:
            for header in (column.key, *column.aliases):
                if header in seen:
                    raise RuntimeError(
                        f"Dwuznaczny nagłówek „{header}” w kontrakcie kolumn {name}"
                    )
                seen.add(header)
                if fold(header) != header:
                    raise RuntimeError(
                        f"Nagłówek „{header}” musi być zapisany bez polskich znaków "
                        "i wielkich liter (porównanie idzie po `fold`)"
                    )


_assert_columns()


# --- Czytanie pliku ----------------------------------------------------

def norm_header(raw: str) -> str:
    """Nagłówek z pliku → postać porównywalna z kontraktem: bez wielkości
    liter i polskich znaków, spacje i myślniki jako podkreślenia."""
    folded = fold(raw or "").strip()
    for char in (" ", "-", ".", "/"):
        folded = folded.replace(char, "_")
    while "__" in folded:
        folded = folded.replace("__", "_")
    return folded.strip("_")


def _cell(value: Any) -> str:
    """Komórka arkusza → tekst. Liczba całkowita zapisana przez Excel jako
    `3.0` wraca jako „3” — inaczej „3.0 serie” trafiłoby do planu."""
    if value is None:
        return ""
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return str(value).strip()


def _read_csv(raw: bytes) -> tuple[list[str], list[list[str]]]:
    for encoding in ("utf-8-sig", "cp1250"):
        try:
            text = raw.decode(encoding)
            break
        except UnicodeDecodeError:
            continue
    else:
        raise SheetError(
            "Nie udało się odczytać pliku. Zapisz go w kodowaniu UTF-8 "
            "(w Excelu: „CSV UTF-8”) albo prześlij jako .xlsx."
        )
    if not text.strip():
        raise SheetError("Plik jest pusty.")
    first_line = text.splitlines()[0]
    delimiter = max((";", ",", "\t"), key=first_line.count)
    if first_line.count(delimiter) == 0:
        delimiter = ","
    reader = csv.reader(io.StringIO(text), delimiter=delimiter)
    # Ten sam limit co w .xlsx. CSV jest juz ograniczony przez MAX_BYTES,
    # wiec nie chodzi tu o bombe, tylko o to, zeby 5 MB tekstu nie robilo
    # sie 50 MB obiektow Pythona, gdy i tak wezmiemy 2000 wierszy.
    rows = []
    for surowy in reader:
        rows.append([_cell(cell) for cell in surowy])
        if len(rows) > MAX_ROWS + 1:
            break
    if not rows:
        raise SheetError("Plik jest pusty.")
    return rows[0], rows[1:]


#: Ile arkusz moze wazyc PO ROZPAKOWANIU. `.xlsx` to archiwum zip, wiec
#: limit MAX_BYTES mierzy plik sprzed rozpakowania i nie mowi nic o tym, co
#: z niego wyjdzie. Prawdziwy arkusz trenera (2000 wierszy) miesci sie tu
#: z ogromnym zapasem; bomba z pomiaru 18.08.2026 mial 423 MB przy 1,64 MB
#: uploadu. Wartosc celowo hojna — chodzi o odciecie absurdu, nie o ciasny
#: limit, ktory odrzuci czyjas duza, ale uczciwa baze.
MAX_ROZPAKOWANE = 100 * 1024 * 1024


def _sprawdz_rozmiar_po_rozpakowaniu(raw: bytes) -> None:
    """Odrzuca bombe dekompresyjna ZANIM openpyxl dotknie pliku.

    DLACZEGO TU, A NIE PRZY CZYTANIU WIERSZY. Przerwanie iteracji na
    MAX_ROWS zbilo czas ze 129 s do 27,5 s i pamiec z 1164 MB do 315 MB,
    ale zmierzenie skladnikow pokazalo, ze reszta idzie na SAM
    `load_workbook`: 24,5 s i 281 MB, zanim odczytamy pierwszy wiersz.
    Iteracja kosztowala juz tylko 0,1 s. Zadne ograniczanie odczytu tego
    nie ruszy — trzeba nie otwierac takiego pliku w ogole.

    Czytamy wylacznie centralny katalog archiwum (`infolist`), czyli
    zadeklarowane rozmiary. Nic nie jest rozpakowywane, wiec sprawdzenie
    jest tanie niezaleznie od tego, co siedzi w srodku.
    """
    import zipfile

    try:
        with zipfile.ZipFile(io.BytesIO(raw)) as archiwum:
            po_rozpakowaniu = sum(wpis.file_size for wpis in archiwum.infolist())
    except zipfile.BadZipFile:
        return  # nie-zip zajmie sie openpyxl i wyda swoj komunikat

    if po_rozpakowaniu > MAX_ROZPAKOWANE:
        raise SheetError(
            f"Arkusz po rozpakowaniu ma {po_rozpakowaniu // (1024 * 1024)} MB "
            f"(limit {MAX_ROZPAKOWANE // (1024 * 1024)} MB) — plik jest za duzy,\n"
            "nawet jesli sam w sobie wyglada na maly. Podziel baze na czesci "
            "albo przeslij jako CSV."
        )


def _read_xlsx(raw: bytes) -> tuple[list[str], list[list[str]]]:
    _sprawdz_rozmiar_po_rozpakowaniu(raw)
    try:
        import openpyxl
    except ImportError:  # pragma: no cover - openpyxl jest zależnością twardą
        raise SheetError("Obsługa .xlsx jest niedostępna; prześlij plik CSV.") from None
    try:
        book = openpyxl.load_workbook(
            io.BytesIO(raw), read_only=True, data_only=True, keep_links=False
        )
    except Exception as exc:  # dowolna awaria parsera = zły plik, nie awaria aplikacji
        raise SheetError(f"Nie udało się otworzyć arkusza: {exc}") from exc
    try:
        sheet = book.worksheets[0]
        # NIE materializujemy calego arkusza. `.xlsx` to archiwum zip:
        # plik 1,64 MB przechodzacy limit 5 MB rozpakowuje sie do 423 MB
        # (3 mln wierszy). MAX_ROWS przycinal WYNIK, gdy wszystko bylo juz
        # w pamieci — zmierzone 18.08.2026: 1164 MB RSS i 129 s na jedno
        # zadanie. Patrz R-19 w rejestrze ryzyk.
        #
        # Czytamy naglowek + MAX_ROWS niepustych wierszy + JEDEN nadmiarowy.
        # Ten jeden jest potrzebny, zeby `read_table` nadal umialo powiedziec
        # „plik ma wiecej niz N wierszy"; bez niego ostrzezenie zniknieloby.
        # Puste wiersze odsiewamy W TRAKCIE, nie po — inaczej arkusz zlozony
        # z miliona pustych wierszy nadal wciagalby cala pamiec.
        rows = []
        for surowy in sheet.iter_rows(values_only=True):
            wiersz = [_cell(cell) for cell in surowy]
            if not any(wiersz):
                continue
            rows.append(wiersz)
            if len(rows) > MAX_ROWS + 1:  # naglowek + MAX_ROWS + nadmiarowy
                break
    finally:
        book.close()
    if not rows:
        raise SheetError("Arkusz jest pusty.")
    return rows[0], rows[1:]


def read_table(
    filename: str, raw: bytes, columns: tuple[Column, ...]
) -> tuple[list[dict[str, str]], list[str], list[str]]:
    """Plik → (wiersze jako słowniki kanoniczne, nieznane nagłówki, ostrzeżenia).

    Wiersz jest słownikiem `kanoniczna_kolumna -> tekst`; kolumny nieobecne
    w pliku po prostu nie występują. Format rozpoznajemy po rozszerzeniu
    nazwy pliku, a nie po zawartości — trener wie, co przesyła, a zgadywanie
    formatu binarnego to proszenie się o cichy błąd."""
    if len(raw) > MAX_BYTES:
        raise SheetError(
            f"Plik jest większy niż {MAX_BYTES // (1024 * 1024)} MB. "
            "Podziel bazę na części albo prześlij jako CSV."
        )
    lowered = (filename or "").lower()
    if lowered.endswith(XLSX_SUFFIXES):
        header_row, data_rows = _read_xlsx(raw)
    elif lowered.endswith(CSV_SUFFIXES) or not lowered:
        header_row, data_rows = _read_csv(raw)
    else:
        raise SheetError(
            "Obsługiwane formaty to CSV (.csv) i arkusz Excela (.xlsx). "
            f"Przesłany plik: „{filename}”."
        )

    index = _index(columns)
    headers: list[str] = []
    unknown: list[str] = []
    warnings: list[str] = []
    for raw_header in header_row:
        normalized = norm_header(raw_header)
        canonical = index.get(normalized)
        if canonical is None:
            headers.append("")
            if normalized:
                unknown.append(raw_header.strip())
            continue
        if canonical in headers:
            warnings.append(
                f"Kolumna „{raw_header.strip()}” powtarza się — brana jest pierwsza."
            )
            headers.append("")
            continue
        headers.append(canonical)

    missing = [c.key for c in columns if c.required and c.key not in headers]
    if missing:
        raise SheetError(
            "Brak wymaganych kolumn: " + ", ".join(missing) + ". "
            "Oczekiwany nagłówek: " + ", ".join(c.key for c in columns) + "."
        )

    rows: list[dict[str, str]] = []
    for row in data_rows:
        if len(rows) >= MAX_ROWS:
            warnings.append(
                f"Plik ma więcej niż {MAX_ROWS} wierszy — reszta została pominięta."
            )
            break
        record = {
            header: (row[position] if position < len(row) else "")
            for position, header in enumerate(headers)
            if header
        }
        if not any(value.strip() for value in record.values()):
            continue  # pusty wiersz — cicho pomijamy
        rows.append(record)
    if not rows:
        raise SheetError("Plik nie zawiera ani jednego wiersza z danymi.")
    return rows, unknown, warnings


# --- Raport ------------------------------------------------------------

@dataclass
class SheetReport:
    """Wynik importu — ten sam kształt dla próby i dla zapisu.

    Rozdzielone celowo: `errors` to wiersze, których NIE zaimportowano
    (trener musi je poprawić), `warnings` to rzeczy zaimportowane, ale
    warte spojrzenia (np. ćwiczenie w szablonie bez odpowiednika w bazie).
    Zlanie ich w jedną listę ukryłoby, co realnie wpadło do bazy."""

    kind: str
    dry_run: bool
    mode: str
    source_ref: str = ""
    rows_read: int = 0
    created: int = 0
    updated: int = 0
    unchanged: int = 0
    skipped: int = 0
    linked: int = 0
    errors: list[dict[str, Any]] = field(default_factory=list)
    warnings: list[str] = field(default_factory=list)
    unknown_columns: list[str] = field(default_factory=list)
    unmapped_muscles: list[str] = field(default_factory=list)
    unlinked_exercises: list[str] = field(default_factory=list)
    created_names: list[str] = field(default_factory=list)
    updated_names: list[str] = field(default_factory=list)
    #: Stan SPRZED importu tych pozycji, których import dotknął — materiał
    #: na punkt przywracania. Wypełniany wyłącznie przy realnym zapisie i
    #: NIE wychodzi w `as_dict()`: to stan wewnętrzny, nie treść odpowiedzi.
    snapshot: list[dict[str, Any]] = field(default_factory=list)

    def error(self, row_no: int, column: str, message: str) -> None:
        self.errors.append({"row": row_no, "column": column, "message": message})
        self.skipped += 1

    def as_dict(self) -> dict[str, Any]:
        return {
            "kind": self.kind,
            "dry_run": self.dry_run,
            "mode": self.mode,
            "source_ref": self.source_ref,
            "rows_read": self.rows_read,
            "created": self.created,
            "updated": self.updated,
            "unchanged": self.unchanged,
            "skipped": self.skipped,
            "linked": self.linked,
            "errors": self.errors,
            "warnings": self.warnings,
            "unknown_columns": self.unknown_columns,
            "unmapped_muscles": self.unmapped_muscles,
            "unlinked_exercises": self.unlinked_exercises,
            # Podgląd, a nie pełna lista — raport ma się zmieścić na ekranie.
            "created_names": self.created_names[:20],
            "updated_names": self.updated_names[:20],
        }


#: Tryby importu do istniejących pozycji.
MODE_FILL = "UZUPELNIJ"
MODE_REPLACE = "ZASTAP"
MODES = (MODE_FILL, MODE_REPLACE)


def normalize_name(text: str) -> str:
    """Nazwa sprowadzona do postaci porównywalnej (ten sam klucz co w
    `import_exercises.normalize_name`) — po nim rozpoznajemy, że pozycja
    już jest w bazie trenera."""
    return " ".join(fold(text).split())


# --- Słowniki: wartość z pliku → klucz kontraktu -----------------------

def _dictionary(keys: tuple[str, ...], labels: dict[str, str]) -> dict[str, str]:
    """Słownik akceptowanych zapisów: klucz („PRZYSIAD”) i polska etykieta
    („przysiad”), oba porównywane przez `fold`. To nie jest zgadywanie —
    to wypisany, zamknięty zbiór dopuszczalnych wartości."""
    mapping: dict[str, str] = {}
    for key in keys:
        mapping[fold(key)] = key
        label = labels.get(key)
        if label:
            mapping[fold(label)] = key
            mapping[fold(label).replace(" ", "_")] = key
    return mapping


GROUP_VALUES = _dictionary(MUSCLE_GROUPS, {g: g.replace("_", " ").lower() for g in MUSCLE_GROUPS})
LEVEL_VALUES = _dictionary(EXERCISE_LEVELS, LEVEL_LABELS)
PATTERN_VALUES = _dictionary(MOVEMENT_PATTERNS, PATTERN_LABELS)
MUSCLE_VALUES = _dictionary(tuple(MUSCLE_LABELS), MUSCLE_LABELS)


def map_value(raw: str, values: dict[str, str]) -> str | None:
    """Wartość z komórki → klucz kontraktu albo None (do raportu)."""
    return values.get(fold(raw or "").strip().replace(" ", "_")) or values.get(
        fold(raw or "").strip()
    )


def split_list(raw: str) -> list[str]:
    """Komórka „a|b|c” → lista. Puste elementy odpadają."""
    return [part.strip() for part in (raw or "").split(LIST_SEPARATOR) if part.strip()]


def map_muscle_cell(raw: str) -> tuple[list[str], list[str]]:
    """Komórka mięśni → (klucze, nierozpoznane zapisy).

    Najpierw próbujemy zamkniętego słownika (klucz lub etykieta), a dopiero
    potem słownika synonimów anatomicznych z parsera opisów — dzięki temu
    „KLATKA_PIERSIOWA”, „klatka piersiowa” i „mięsień piersiowy większy”
    trafiają w to samo miejsce, a nazwa zbiorcza („góra ciała”) nadal
    zostaje nierozpoznana i idzie do raportu."""
    keys: list[str] = []
    unmapped: list[str] = []
    for part in (raw or "").split(MUSCLE_SEPARATOR):
        phrase = part.strip()
        if not phrase:
            continue
        direct = map_value(phrase, MUSCLE_VALUES)
        hits = [direct] if direct else map_muscle_phrase(phrase)
        if not hits:
            unmapped.append(phrase)
            continue
        for key in hits:
            if key not in keys:
                keys.append(key)
    return keys, unmapped


# --- Import ćwiczeń ----------------------------------------------------

#: Pola tekstowe ćwiczenia: kolumna pliku → atrybut modelu.
_EXERCISE_TEXT_FIELDS: tuple[tuple[str, str], ...] = (
    ("opis", "how_to"),
    ("nazwa_en", "name_en"),
    ("efekt", "benefit"),
    ("sprzet", "equipment"),
    ("bezpieczenstwo", "safety"),
    ("latwiej", "easier"),
    ("trudniej", "harder"),
    ("tempo", "tempo_hint"),
    ("oddech", "breathing"),
    ("wideo_url", "video_url"),
    ("zrodlo", "source_ref"),
)
#: Pola listowe: kolumna pliku → atrybut modelu (JSON w kolumnie tekstowej).
_EXERCISE_LIST_FIELDS: tuple[tuple[str, str], ...] = (
    ("kroki", "steps_json"),
    ("bledy", "mistakes_json"),
    ("wskazowki", "cues_json"),
    ("tagi", "tags_json"),
)


def _changes_for(
    target: Exercise | None, values: dict[str, Any], *, mode: str
) -> dict[str, Any]:
    """Wartości z pliku → realne zmiany na istniejącej pozycji.

    Pusta wartość NIGDY nie kasuje tego, co jest już w bazie — także w
    trybie `ZASTAP`. Brak komórki w pliku znaczy „nie mam tej informacji”,
    a nie „usuń”; skasowanie opisu przez pominiętą kolumnę byłoby cichą
    utratą pracy trenera. W trybie `UZUPELNIJ` zmieniamy wyłącznie pola,
    które w bazie są puste."""
    changes: dict[str, Any] = {}
    for attr, value in values.items():
        if value in (None, "", []):
            continue
        current = getattr(target, attr, None) if target is not None else None
        if current and mode == MODE_FILL:
            continue
        if current == value:
            continue
        changes[attr] = value
    return changes


def import_exercises_sheet(
    db: Session,
    coach_id: str,
    rows: list[dict[str, str]],
    *,
    mode: str = MODE_FILL,
    dry_run: bool = True,
    source_ref: str = "",
) -> SheetReport:
    """Wiersze arkusza → baza ćwiczeń TEGO trenera.

    Dopasowanie do istniejącej pozycji idzie po znormalizowanej nazwie i
    obejmuje także ćwiczenia zarchiwizowane — inaczej import robiłby
    duplikat czegoś, co trener świadomie schował.

    Przy `dry_run=True` funkcja NIE dotyka ani jednego obiektu sesji (ten
    sam wzorzec co `import_exercises.import_library`), więc podgląd da się
    bezpiecznie uruchomić w dowolnym momencie. Nie commituje — o trwałości
    decyduje wywołujący."""
    if mode not in MODES:
        raise SheetError(f"Nieznany tryb importu: {mode}. Dopuszczalne: {', '.join(MODES)}.")
    report = SheetReport(kind="EXERCISES", dry_run=dry_run, mode=mode,
                         source_ref=source_ref, rows_read=len(rows))
    existing: dict[str, Exercise] = {}
    for row in db.query(Exercise).filter(Exercise.coach_id == coach_id).all():
        existing.setdefault(normalize_name(row.name), row)

    seen: set[str] = set()
    unmapped: list[str] = []
    row_no = 1  # wiersz 1 = nagłówek

    for record in rows:
        row_no += 1
        name = record.get("nazwa", "").strip()
        if not name:
            report.error(row_no, "nazwa", "nazwa ćwiczenia jest wymagana")
            continue
        if len(name) > 300:
            report.error(row_no, "nazwa", "nazwa dłuższa niż 300 znaków")
            continue
        key = normalize_name(name)
        if key in seen:
            report.error(row_no, "nazwa", f"„{name}” powtarza się w pliku — wiersz pominięty")
            continue
        seen.add(key)

        target = existing.get(key)
        is_new = target is None

        group_raw = record.get("grupa", "").strip()
        group = map_value(group_raw, GROUP_VALUES) if group_raw else None
        if group_raw and group is None:
            report.error(
                row_no, "grupa",
                f"„{group_raw}” to nie jest grupa ze słownika. Dopuszczalne: "
                + ", ".join(MUSCLE_GROUPS),
            )
            continue
        how_to = record.get("opis", "").strip()
        if is_new and group is None:
            report.error(row_no, "grupa", "nowe ćwiczenie musi mieć grupę mięśniową")
            continue
        if is_new and not how_to:
            report.error(row_no, "opis", "nowe ćwiczenie musi mieć opis wykonania")
            continue

        level = pattern = None
        for column, table, allowed in (
            ("poziom", LEVEL_VALUES, EXERCISE_LEVELS),
            ("wzorzec", PATTERN_VALUES, MOVEMENT_PATTERNS),
        ):
            raw = record.get(column, "").strip()
            if not raw:
                continue
            mapped = map_value(raw, table)
            if mapped is None:
                report.warnings.append(
                    f"Wiersz {row_no}: „{raw}” to nie jest wartość ze słownika "
                    f"({column}) — pole zostało puste. Dopuszczalne: " + ", ".join(allowed)
                )
                continue
            if column == "poziom":
                level = mapped
            else:
                pattern = mapped

        primary, bad_primary = map_muscle_cell(record.get("miesnie_glowne", ""))
        secondary, bad_secondary = map_muscle_cell(record.get("miesnie_pomocnicze", ""))
        secondary = [k for k in secondary if k not in primary]
        for phrase in (*bad_primary, *bad_secondary):
            if phrase not in unmapped:
                unmapped.append(phrase)

        values: dict[str, Any] = {"muscle_group": group}
        for column, attr in _EXERCISE_TEXT_FIELDS:
            # `source_ref` z nazwy pliku dostaje WYŁĄCZNIE nowa pozycja
            # (niżej, przy tworzeniu). Proweniencja mówi, skąd wpis
            # pochodzi — istniejące ćwiczenie nie pochodzi z tego pliku, a
            # doklejanie mu nazwy pliku psułoby też idempotencję: eksport
            # wgrany z powrotem „zmieniałby” każdy wiersz.
            values[attr] = record.get(column, "").strip() or None
        for column, attr in _EXERCISE_LIST_FIELDS:
            items = split_list(record.get(column, ""))
            values[attr] = json.dumps(items, ensure_ascii=False) if items else None
        values["level"] = level
        values["pattern"] = pattern
        values["muscles_primary"] = join_muscles(primary)
        values["muscles_secondary"] = join_muscles(secondary)

        if is_new:
            report.created += 1
            report.created_names.append(name)
            if dry_run:
                continue
            item = Exercise(
                id=new_id("EXC"), coach_id=coach_id, created_by=coach_id,
                name=name, muscle_group=group or "INNE", how_to=how_to,
                source_kind=SOURCE_IMPORTED, status="ACTIVE",
            )
            for attr, value in values.items():
                if attr in ("muscle_group", "how_to") or value in (None, "", []):
                    continue
                setattr(item, attr, value)
            if source_ref and not item.source_ref:
                item.source_ref = source_ref
            db.add(item)
            existing[key] = item
            report.snapshot.append({"id": item.id, "created": True, "before": {}})
            continue

        changes = _changes_for(target, values, mode=mode)
        if not changes:
            report.unchanged += 1
            continue
        report.updated += 1
        report.updated_names.append(name)
        if dry_run:
            continue
        report.snapshot.append({
            "id": target.id, "created": False,
            "before": {attr: getattr(target, attr) for attr in changes},
        })
        for attr, value in changes.items():
            setattr(target, attr, value)
        target.updated_at = now_iso()

    report.unmapped_muscles = unmapped
    return report


# --- Import szablonów treningowych -------------------------------------

#: Pola pozycji planu: kolumna pliku → klucz w treści wersji (`ExerciseIn`).
_TEMPLATE_ITEM_FIELDS: tuple[tuple[str, str], ...] = (
    ("serie", "sets"),
    ("powtorzenia", "reps"),
    ("ciezar", "weight"),
    ("tempo", "tempo"),
    ("przerwa", "rest"),
    ("komentarz", "comment"),
    ("wideo_url", "video_url"),
)
#: Limity zgodne z `schemas.ExerciseIn` — plik nie może wprowadzić do bazy
#: treści, której nie przyjąłby zwykły formularz.
_ITEM_LIMITS: dict[str, int] = {
    "sets": 40, "reps": 40, "weight": 40, "tempo": 40, "rest": 40,
    "comment": 1000, "video_url": 500,
}
MAX_DAYS = 14
MAX_ITEMS_PER_DAY = 40


def _parse_int(raw: str) -> int | None:
    try:
        return int(float((raw or "").strip().replace(",", ".")))
    except (TypeError, ValueError):
        return None


def _template_content(days: list[dict[str, Any]]) -> str:
    return json.dumps({"days": days}, ensure_ascii=False)


def import_templates_sheet(
    db: Session,
    coach_id: str,
    rows: list[dict[str, str]],
    *,
    dry_run: bool = True,
    source_ref: str = "",
) -> SheetReport:
    """Wiersze arkusza → szablony treningowe TEGO trenera.

    Jeden wiersz to jedno ćwiczenie w jednym dniu jednego szablonu; wiersze
    grupuje kolumna `szablon`. Kolejność dni i pozycji bierzemy z kolumn
    `dzien_nr`/`pozycja`, a gdy ich nie ma — z kolejności w pliku.

    Nazwa ćwiczenia jest DOPASOWYWANA do aktywnej bazy trenera i zapisywana
    jako miękkie odniesienie `exercise_id` (dokładnie ten sam kontrakt, co
    przy ręcznym układaniu planu). Brak dopasowania NIE jest błędem: pozycja
    wchodzi do szablonu z samą nazwą i trafia na listę `unlinked_exercises`,
    żeby trener wiedział, co warto dodać do bazy.

    Szablon o tej samej nazwie nie jest nadpisywany — dostaje NOWĄ WERSJĘ z
    powodem wskazującym plik źródłowy. Wersja o identycznej treści nie
    powstaje w ogóle (idempotencja). Przy `dry_run=True` sesja pozostaje
    nietknięta."""
    report = SheetReport(kind="TEMPLATES", dry_run=dry_run, mode=MODE_FILL,
                         source_ref=source_ref, rows_read=len(rows))

    catalog: dict[str, str] = {}
    for row in (
        db.query(Exercise)
        .filter(Exercise.coach_id == coach_id, Exercise.status == "ACTIVE")
        .all()
    ):
        catalog.setdefault(normalize_name(row.name), row.id)

    # Szablony w kolejności pierwszego wystąpienia w pliku; wewnątrz —
    # dni, wewnątrz dni — pozycje. Zwykłe słowniki, bo od 3.7 zachowują
    # kolejność wstawiania, a ona jest tu częścią kontraktu.
    plans: dict[str, dict[str, Any]] = {}
    unlinked: list[str] = []
    row_no = 1

    for record in rows:
        row_no += 1
        title = record.get("szablon", "").strip()
        day_name = record.get("dzien", "").strip()
        name = record.get("cwiczenie", "").strip()
        if not title:
            report.error(row_no, "szablon", "nazwa szablonu jest wymagana")
            continue
        if not day_name:
            report.error(row_no, "dzien", "nazwa dnia jest wymagana")
            continue
        if not name:
            report.error(row_no, "cwiczenie", "nazwa ćwiczenia jest wymagana")
            continue
        if len(title) > 300 or len(day_name) > 200 or len(name) > 300:
            report.error(row_no, "-", "nazwa przekracza dopuszczalną długość")
            continue

        plan = plans.setdefault(title, {"days": {}})
        day = plan["days"].get(day_name)
        if day is None:
            if len(plan["days"]) >= MAX_DAYS:
                report.error(
                    row_no, "dzien",
                    f"szablon „{title}” ma już {MAX_DAYS} dni — wiersz pominięty",
                )
                continue
            day = {"name": day_name, "weekday": None, "order": None, "exercises": []}
            plan["days"][day_name] = day
        if len(day["exercises"]) >= MAX_ITEMS_PER_DAY:
            report.error(
                row_no, "cwiczenie",
                f"dzień „{day_name}” ma już {MAX_ITEMS_PER_DAY} pozycji — wiersz pominięty",
            )
            continue

        weekday = _parse_int(record.get("dzien_tygodnia", ""))
        if weekday is not None and 1 <= weekday <= 7:
            day["weekday"] = weekday
        elif record.get("dzien_tygodnia", "").strip():
            report.warnings.append(
                f"Wiersz {row_no}: dzień tygodnia „{record['dzien_tygodnia']}” "
                "poza zakresem 1–7 — pominięty."
            )
        day_order = _parse_int(record.get("dzien_nr", ""))
        if day_order is not None and day["order"] is None:
            day["order"] = day_order

        item: dict[str, Any] = {"name": name, "exercise_id": catalog.get(normalize_name(name))}
        if item["exercise_id"]:
            report.linked += 1
        elif name not in unlinked:
            unlinked.append(name)
        for column, key in _TEMPLATE_ITEM_FIELDS:
            value = record.get(column, "").strip()
            limit = _ITEM_LIMITS[key]
            if value and len(value) > limit:
                report.warnings.append(
                    f"Wiersz {row_no}: „{column}” dłuższe niż {limit} znaków — przycięte."
                )
                value = value[:limit]
            item[key] = value or None
        item["order"] = _parse_int(record.get("pozycja", ""))
        day["exercises"].append(item)

    report.unlinked_exercises = unlinked

    existing: dict[str, TrainingPlan] = {}
    for plan_row in (
        db.query(TrainingPlan)
        .filter(
            TrainingPlan.coach_id == coach_id,
            TrainingPlan.is_template.is_(True),
            TrainingPlan.status == "ACTIVE",
        )
        .all()
    ):
        existing.setdefault(normalize_name(plan_row.title), plan_row)

    reason = f"Import z pliku: {source_ref}" if source_ref else "Import z pliku"
    for title, plan in plans.items():
        days_out: list[dict[str, Any]] = []
        ordered_days = sorted(
            plan["days"].values(),
            key=lambda d, order=list(plan["days"]): (
                d["order"] if d["order"] is not None else 10_000 + order.index(d["name"])
            ),
        )
        for day in ordered_days:
            items = sorted(
                day["exercises"],
                key=lambda e, seq=day["exercises"]: (
                    e["order"] if e["order"] is not None else 10_000 + seq.index(e)
                ),
            )
            days_out.append({
                "name": day["name"],
                "weekday": day["weekday"],
                "exercises": [{k: v for k, v in item.items() if k != "order"} for item in items],
            })
        content = _template_content(days_out)

        current = existing.get(normalize_name(title))
        if current is None:
            report.created += 1
            report.created_names.append(title)
            if dry_run:
                continue
            plan_row = TrainingPlan(
                id=new_id("PLN"), client_id=None, coach_id=coach_id, title=title,
                status="ACTIVE", current_version_no=1, is_template=True,
            )
            db.add(plan_row)
            db.add(TrainingPlanVersion(
                id=new_id("PLV"), plan_id=plan_row.id, version_no=1,
                reason=reason, content_json=content, created_by=coach_id,
            ))
            report.snapshot.append({"id": plan_row.id, "created": True, "version_no": 0})
            continue

        latest = (
            db.query(TrainingPlanVersion)
            .filter_by(plan_id=current.id, version_no=current.current_version_no)
            .one_or_none()
        )
        if latest is not None and latest.content_json == content:
            report.unchanged += 1
            continue
        report.updated += 1
        report.updated_names.append(title)
        if dry_run:
            continue
        report.snapshot.append({
            "id": current.id, "created": False,
            "version_no": current.current_version_no,
        })
        current.current_version_no += 1
        current.updated_at = now_iso()
        db.add(TrainingPlanVersion(
            id=new_id("PLV"), plan_id=current.id, version_no=current.current_version_no,
            reason=reason, content_json=content, created_by=coach_id,
        ))

    return report


# --- Wzór pliku i eksport (podróż w obie strony) -----------------------

def schema_dict(columns: tuple[Column, ...]) -> list[dict[str, Any]]:
    """Kontrakt kolumn dla interfejsu — ta sama definicja, z której
    powstaje wzór pliku i dokumentacja. Jedno źródło prawdy."""
    return [
        {
            "key": column.key, "label": column.label, "required": column.required,
            "example": column.example, "aliases": list(column.aliases),
        }
        for column in columns
    ]


def dictionaries() -> dict[str, list[dict[str, str]]]:
    """Zamknięte słowniki wraz z polskimi etykietami — trener ma widzieć
    dopuszczalne wartości w aplikacji, a nie szukać ich w dokumentacji."""
    return {
        "grupa": [{"key": g, "label": g.replace("_", " ").lower()} for g in MUSCLE_GROUPS],
        "poziom": [{"key": k, "label": LEVEL_LABELS[k]} for k in EXERCISE_LEVELS],
        "wzorzec": [{"key": k, "label": PATTERN_LABELS[k]} for k in MOVEMENT_PATTERNS],
        "miesnie": [{"key": k, "label": v} for k, v in MUSCLE_LABELS.items()],
    }


def _csv_bytes(header: list[str], rows: list[list[str]]) -> bytes:
    buffer = io.StringIO()
    writer = csv.writer(buffer, delimiter=",", lineterminator="\n")
    writer.writerow(header)
    for row in rows:
        writer.writerow(row)
    # BOM: arkusze kalkulacyjne otwierają wtedy polskie znaki poprawnie
    # (ten sam zabieg co przy eksporcie bazy produktów).
    return ("﻿" + buffer.getvalue()).encode("utf-8")


def example_csv(columns: tuple[Column, ...]) -> bytes:
    """Wzór do pobrania: nagłówek zgodny z kontraktem i JEDEN wiersz
    przykładowy. Wzór ma pokazywać format, a nie podsuwać treść — dlatego
    przykład jest jeden i oczywisty do skasowania."""
    return _csv_bytes([c.key for c in columns], [[c.example for c in columns]])


def exercises_csv(items: list[Exercise]) -> bytes:
    """Eksport bazy ćwiczeń w DOKŁADNIE tym formacie, który przyjmuje
    import — prawo wyjścia i zarazem najwygodniejszy sposób masowej
    edycji: pobierz, popraw w arkuszu, wgraj z powrotem."""
    def joined(raw: str | None) -> str:
        if not raw:
            return ""
        try:
            return LIST_SEPARATOR.join(json.loads(raw))
        except (ValueError, TypeError):
            return ""

    rows: list[list[str]] = []
    for item in items:
        mapping = {
            "nazwa": item.name, "grupa": item.muscle_group, "opis": item.how_to,
            "nazwa_en": item.name_en, "efekt": item.benefit, "sprzet": item.equipment,
            "poziom": item.level, "wzorzec": item.pattern,
            "miesnie_glowne": item.muscles_primary,
            "miesnie_pomocnicze": item.muscles_secondary,
            "kroki": joined(item.steps_json), "bledy": joined(item.mistakes_json),
            "wskazowki": joined(item.cues_json), "bezpieczenstwo": item.safety,
            "latwiej": item.easier, "trudniej": item.harder, "tempo": item.tempo_hint,
            "oddech": item.breathing, "tagi": joined(item.tags_json),
            "wideo_url": item.video_url, "zrodlo": item.source_ref,
        }
        rows.append([str(mapping.get(c.key) or "") for c in EXERCISE_COLUMNS])
    return _csv_bytes([c.key for c in EXERCISE_COLUMNS], rows)


def templates_csv(plans: list[tuple[TrainingPlan, str]]) -> bytes:
    """Eksport szablonów w formacie importu: jeden wiersz = jedna pozycja.
    `plans` to pary (szablon, treść bieżącej wersji jako JSON)."""
    rows: list[list[str]] = []
    for plan, content_json in plans:
        try:
            days = json.loads(content_json).get("days", [])
        except (ValueError, TypeError, AttributeError):
            continue
        for day_no, day in enumerate(days, start=1):
            for position, item in enumerate(day.get("exercises", []), start=1):
                mapping = {
                    "szablon": plan.title, "dzien": day.get("name", ""),
                    "dzien_nr": day_no, "dzien_tygodnia": day.get("weekday"),
                    "pozycja": position, "cwiczenie": item.get("name", ""),
                    "serie": item.get("sets"), "powtorzenia": item.get("reps"),
                    "ciezar": item.get("weight"), "tempo": item.get("tempo"),
                    "przerwa": item.get("rest"), "komentarz": item.get("comment"),
                    "wideo_url": item.get("video_url"),
                }
                rows.append([str(mapping.get(c.key) or "") for c in TEMPLATE_COLUMNS])
    return _csv_bytes([c.key for c in TEMPLATE_COLUMNS], rows)


# --- Punkt przywracania: „cofnij ten import" --------------------------

#: Pola ćwiczenia, które import może zmienić — dokładnie te i tylko te
#: trafiają do migawki i tylko te przywraca cofnięcie. Lista wyprowadzona
#: z kontraktu kolumn, żeby nowa kolumna importu nie mogła po cichu
#: wypaść poza zasięg cofania.
SNAPSHOT_FIELDS: tuple[str, ...] = (
    "muscle_group", "level", "pattern", "muscles_primary", "muscles_secondary",
    *(attr for _, attr in _EXERCISE_TEXT_FIELDS),
    *(attr for _, attr in _EXERCISE_LIST_FIELDS),
)

#: Ile migawek trzymamy na trenera. Cofnięcie ma sens tuż po imporcie —
#: starsze i tak przywracałyby stan sprzed późniejszych, świadomych zmian.
SNAPSHOT_KEEP = 20


def _assert_snapshot_covers_import() -> None:
    """Każde pole, które import ustawia, musi być w migawce. Sprawdzane
    przy imporcie modułu — inaczej dołożenie kolumny do importu po cichu
    wyłączyłoby dla niej cofanie."""
    written = {"muscle_group", "how_to", "level", "pattern",
               "muscles_primary", "muscles_secondary"}
    written |= {attr for _, attr in _EXERCISE_TEXT_FIELDS}
    written |= {attr for _, attr in _EXERCISE_LIST_FIELDS}
    missing = written - set(SNAPSHOT_FIELDS)
    if missing:
        raise RuntimeError(
            "Pola zapisywane przez import, których nie obejmuje migawka "
            f"(nie dałoby się ich cofnąć): {sorted(missing)}"
        )


_assert_snapshot_covers_import()


def store_snapshot(db: Session, coach_id: str, report: SheetReport) -> str | None:
    """Zapisuje punkt przywracania dla właśnie wykonanego importu.

    Zwraca identyfikator migawki albo None, gdy nie ma czego cofać (import
    niczego nie zmienił). Stare migawki są przycinane do `SNAPSHOT_KEEP` —
    cofnięcie sprzed wielu operacji przywracałoby stan sprzed świadomych,
    późniejszych zmian trenera, o których ta migawka nic nie wie."""
    if report.dry_run or not report.snapshot:
        return None
    row = ImportSnapshot(
        id=new_id("IMS"), coach_id=coach_id, kind=report.kind,
        source_ref=report.source_ref or "plik", mode=report.mode,
        rows=len(report.snapshot),
        payload_json=json.dumps(report.snapshot, ensure_ascii=False),
    )
    db.add(row)
    db.flush()
    stale = (
        db.query(ImportSnapshot)
        .filter(ImportSnapshot.coach_id == coach_id)
        .order_by(ImportSnapshot.created_at.desc())
        .offset(SNAPSHOT_KEEP)
        .all()
    )
    for old in stale:
        db.delete(old)
    return row.id


def snapshot_out(row: ImportSnapshot) -> dict[str, Any]:
    return {
        "id": row.id, "kind": row.kind, "source_ref": row.source_ref,
        "mode": row.mode, "rows": row.rows, "created_at": row.created_at,
        "restored_at": row.restored_at,
    }


def undo_import(db: Session, coach_id: str, row: ImportSnapshot) -> dict[str, Any]:
    """Cofa jeden import do stanu sprzed niego.

    DWIE REGUŁY, które odróżniają cofnięcie od kasowania:

    * pozycja UTWORZONA przez import zostaje **zarchiwizowana**, nigdy
      usunięta — historia zostaje, a trener może ją przywrócić ręcznie;
    * pozycja ZMIENIONA wraca do wartości sprzed importu, pole po polu, i
      wyłącznie w polach, których import dotknął. Szablon nie cofa się
      przez skasowanie wersji: dostaje **nową wersję** z treścią sprzed
      importu, więc pełna historia (łącznie z samym importem) zostaje.

    Cofnięcie jest jednorazowe — patrz `ImportSnapshot.restored_at`."""
    if row.restored_at:
        raise SheetError("Ten import został już cofnięty.")
    payload = json.loads(row.payload_json)
    restored = archived = missing = 0

    if row.kind == "EXERCISES":
        for entry in payload:
            item = db.get(Exercise, entry["id"])
            if item is None or item.coach_id != coach_id:
                missing += 1
                continue
            if entry.get("created"):
                item.status = "ARCHIVED"
                item.updated_at = now_iso()
                archived += 1
                continue
            for attr, value in entry.get("before", {}).items():
                if attr in SNAPSHOT_FIELDS:
                    setattr(item, attr, value)
            item.updated_at = now_iso()
            restored += 1
    else:
        for entry in payload:
            plan = db.get(TrainingPlan, entry["id"])
            if plan is None or plan.coach_id != coach_id:
                missing += 1
                continue
            if entry.get("created"):
                plan.status = "ARCHIVED"
                plan.updated_at = now_iso()
                archived += 1
                continue
            previous = (
                db.query(TrainingPlanVersion)
                .filter_by(plan_id=plan.id, version_no=entry["version_no"])
                .one_or_none()
            )
            if previous is None:
                missing += 1
                continue
            plan.current_version_no += 1
            plan.updated_at = now_iso()
            db.add(TrainingPlanVersion(
                id=new_id("PLV"), plan_id=plan.id,
                version_no=plan.current_version_no,
                reason=f"Cofnięcie importu z pliku: {row.source_ref}",
                content_json=previous.content_json, created_by=coach_id,
            ))
            restored += 1

    row.restored_at = now_iso()
    return {"restored": restored, "archived": archived, "missing": missing}
