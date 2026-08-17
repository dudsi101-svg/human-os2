"""Monitoring w czasie: adherencja harmonogramu, dziennik obserwacji
(nigdy diagnoza — wyłącznie flaga do przeglądu przez trenera), dziennik
żywieniowy, agregacja trendów; eksport Excel."""

import io

from conftest import CLIENT_A, COACH, get_user_id, login


def _first_supplement_item(seeded, hc, id_a):
    items = seeded.get(f"/api/clients/{id_a}/schedule", headers=hc).json()["items"]
    return next(i for i in items if i["category"] == "SUPLEMENT")


def test_schedule_completion_is_idempotent_and_visible_on_today(seeded):
    ha = login(seeded, CLIENT_A)
    hc = login(seeded, COACH)
    id_a = get_user_id(seeded, ha)
    item = _first_supplement_item(seeded, hc, id_a)

    today = seeded.get("/api/me/today", headers=ha).json()["date"]
    r = seeded.post(f"/api/clients/{id_a}/schedule/{item['id']}/complete", headers=ha,
                    json={"completed_on": today})
    assert r.status_code == 201
    # Ponowne wywołanie na ten sam dzień nadpisuje, nie duplikuje.
    r2 = seeded.post(f"/api/clients/{id_a}/schedule/{item['id']}/complete", headers=ha,
                     json={"completed_on": today, "status": "SKIPPED", "note": "zapomniałem"})
    assert r2.status_code == 201
    assert r2.json()["id"] == r.json()["id"]

    monitoring = seeded.get(f"/api/clients/{id_a}/monitoring", headers=ha).json()
    bucket = monitoring["adherence"]["SUPLEMENT"]
    assert bucket["total"] >= 1


def test_other_client_cannot_complete_schedule_item(seeded):
    ha = login(seeded, CLIENT_A)
    hb = login(seeded, {"email": "klient.b@example.com", "password": "KlientB#2026!x"})
    hc = login(seeded, COACH)
    id_a = get_user_id(seeded, ha)
    item = _first_supplement_item(seeded, hc, id_a)
    r = seeded.post(f"/api/clients/{id_a}/schedule/{item['id']}/complete", headers=hb,
                    json={"completed_on": "2026-08-17"})
    assert r.status_code == 404


def test_observation_info_vs_flagged_and_coach_notification(seeded):
    ha = login(seeded, CLIENT_A)
    hc = login(seeded, COACH)
    id_a = get_user_id(seeded, ha)

    r = seeded.post(f"/api/clients/{id_a}/observations", headers=ha, json={
        "occurred_on": "2026-08-17", "category": "SAMOPOCZUCIE",
        "severity": "INFO", "text": "Czuję się dobrze po treningu.",
    })
    assert r.status_code == 201

    item = _first_supplement_item(seeded, hc, id_a)
    r = seeded.post(f"/api/clients/{id_a}/observations", headers=ha, json={
        "occurred_on": "2026-08-17", "schedule_item_id": item["id"],
        "category": "REAKCJA", "severity": "NIEPOKOJACE",
        "text": "Po kreatynie lekkie mdłości.",
    })
    assert r.status_code == 201

    obs = seeded.get(f"/api/clients/{id_a}/observations", headers=hc).json()["observations"]
    assert obs[0]["severity"] == "NIEPOKOJACE"
    assert obs[0]["schedule_item_name"] == item["name"]

    # Panel trenera flaguje niepokojącą obserwację (bez oceniania/diagnozy).
    clients = seeded.get("/api/coach/clients", headers=hc).json()["clients"]
    row = next(c for c in clients if c["client_id"] == id_a)
    assert row["flags"]["flagged_observations"] >= 1


def test_observation_never_auto_diagnoses_only_logs_for_review(seeded):
    """System nie interpretuje obserwacji — zapisuje tekst dosłownie,
    zwraca tylko flagę ważności do przeglądu przez człowieka."""
    ha = login(seeded, CLIENT_A)
    id_a = get_user_id(seeded, ha)
    text = "Ból głowy po suplemencie X, nie wiem czy to normalne."
    r = seeded.post(f"/api/clients/{id_a}/observations", headers=ha, json={
        "occurred_on": "2026-08-17", "category": "OBJAW",
        "severity": "NIEPOKOJACE", "text": text,
    })
    assert r.status_code == 201
    obs = seeded.get(f"/api/clients/{id_a}/observations", headers=ha).json()["observations"]
    stored = next(o for o in obs if o["text"] == text)
    assert stored["text"] == text  # bez modyfikacji/interpretacji
    assert set(stored.keys()) == {
        "id", "occurred_on", "category", "severity", "text",
        "schedule_item_id", "schedule_item_name", "created_by", "created_at",
    }


def test_invalid_schedule_item_rejected_for_observation(seeded):
    ha = login(seeded, CLIENT_A)
    id_a = get_user_id(seeded, ha)
    r = seeded.post(f"/api/clients/{id_a}/observations", headers=ha, json={
        "occurred_on": "2026-08-17", "schedule_item_id": "HOS-SCH-DOESNOTEXIST",
        "category": "INNE", "severity": "INFO", "text": "test",
    })
    assert r.status_code == 422


def test_nutrition_log_upsert_and_monitoring_series(seeded):
    ha = login(seeded, CLIENT_A)
    id_a = get_user_id(seeded, ha)
    r = seeded.post(f"/api/clients/{id_a}/nutrition-log", headers=ha, json={
        "logged_on": "2026-08-17", "kcal": 2100, "protein_g": 170, "water_l": 2.5,
    })
    assert r.status_code == 201
    log_id = r.json()["id"]
    r2 = seeded.post(f"/api/clients/{id_a}/nutrition-log", headers=ha, json={
        "logged_on": "2026-08-17", "kcal": 2250,
    })
    assert r2.json()["id"] == log_id  # upsert, nie duplikat

    logs = seeded.get(f"/api/clients/{id_a}/nutrition-log", headers=ha).json()["logs"]
    entry = next(x for x in logs if x["logged_on"] == "2026-08-17")
    assert entry["kcal"] == 2250

    monitoring = seeded.get(f"/api/clients/{id_a}/monitoring", headers=ha).json()
    assert any(pt["date"] == "2026-08-17" and pt["value"] == 2250
              for pt in monitoring["nutrition"]["log_series"])
    assert monitoring["nutrition"]["target_kcal"] == 2300  # z diety seedu


def test_monitoring_includes_goal_and_measurement_trends(seeded):
    ha = login(seeded, CLIENT_A)
    id_a = get_user_id(seeded, ha)
    data = seeded.get(f"/api/clients/{id_a}/monitoring", headers=ha).json()
    assert data["goal"]["title"]
    assert data["goal"]["days_remaining"] is not None
    assert "weight" in data["measurement_series"]
    assert len(data["measurement_series"]["weight"]) >= 5  # 8 tyg. seedu


def test_monitoring_requires_access(seeded):
    hb = login(seeded, {"email": "klient.b@example.com", "password": "KlientB#2026!x"})
    ha = login(seeded, CLIENT_A)
    id_a = get_user_id(seeded, ha)
    assert seeded.get(f"/api/clients/{id_a}/monitoring", headers=hb).status_code == 404


def test_excel_export_downloads_workbook(seeded):
    ha = login(seeded, CLIENT_A)
    r = seeded.get("/api/me/export.xlsx", headers=ha)
    assert r.status_code == 200
    assert r.headers["content-type"].startswith(
        "application/vnd.openxmlformats-officedocument"
    )
    assert "attachment" in r.headers["content-disposition"]
    from openpyxl import load_workbook

    wb = load_workbook(io.BytesIO(r.content))
    assert "konto" in wb.sheetnames
    assert "measurements" in wb.sheetnames


def test_voice_message_upload_allowed(seeded):
    ha = login(seeded, CLIENT_A)
    audio = b"\x1aE\xdf\xa3" + b"0" * 100  # nagłówek WebM (uproszczony)
    r = seeded.post("/api/files", headers=ha, files={
        "file": ("wiadomosc.webm", io.BytesIO(audio), "audio/webm")})
    assert r.status_code == 201
    assert r.json()["content_type"] == "audio/webm"
